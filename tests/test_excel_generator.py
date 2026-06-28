from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from app.models import ParsedOFXData, SummaryTotals, TransactionRecord
from app.services.excel_generator import generate_excel_report


def test_generate_excel_report(tmp_path: Path) -> None:
    parsed_data = ParsedOFXData(
        original_filename="extrato.ofx",
        banco="Itau",
        processed_at=datetime(2026, 6, 25, 10, 30, 0),
        records=[
            TransactionRecord(
                entrada="1500.00",
                item="PIX RECEBIDO CLIENTE X",
                saida="",
                data="25/06/2026",
                banco="Itau",
            )
        ],
        summary=SummaryTotals(
            total_entradas="1500.00",
            total_saidas="0.00",
            saldo="1500.00",
            quantidade_movimentacoes=1,
        ),
    )

    output = generate_excel_report(tmp_path / "saida.xlsx", parsed_data)
    workbook = load_workbook(output)

    assert output.exists()
    assert "Movimentações" in workbook.sheetnames
    assert "Resumo" in workbook.sheetnames
    assert workbook["Movimentações"]["A1"].value == "Entrada"
    assert workbook["Resumo"]["A1"].value == "Resumo da Conversão"
