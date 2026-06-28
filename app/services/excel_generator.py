from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models import ParsedOFXData


NAVY = "081526"
NAVY_LIGHT = "10243A"
CYAN = "00C8F2"
CYAN_DARK = "009EC3"
WHITE = "FFFFFF"
TEXT = "172033"
BORDER = "D9E2EC"
SUMMARY_BG = "E6F9FE"


def generate_excel_report(output_path: Path, parsed_data: ParsedOFXData) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    movements_sheet = workbook.active
    movements_sheet.title = "Movimentações"
    summary_sheet = workbook.create_sheet("Resumo")

    _build_movements_sheet(movements_sheet, parsed_data)
    _build_summary_sheet(summary_sheet, parsed_data)

    workbook.save(output_path)
    return output_path


def _build_movements_sheet(sheet, parsed_data: ParsedOFXData) -> None:
    headers = ["Entrada", "Item", "Saída", "Data", "Banco"]
    sheet.append(headers)
    _style_header_row(sheet, 1)

    currency_format = 'R$ #,##0.00'
    date_format = "dd/mm/yyyy"

    for record in parsed_data.records:
        sheet.append(
            [
                _to_float(record.entrada),
                record.item,
                _to_float(record.saida),
                _to_excel_date(record.data),
                record.banco,
            ]
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    thin_cyan = Side(style="thin", color=CYAN)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.border = Border(bottom=thin_cyan)
            cell.alignment = Alignment(vertical="center")
            if cell.column in (1, 3) and cell.value is not None:
                cell.number_format = currency_format
            if cell.column == 4 and cell.value is not None:
                cell.number_format = date_format

    _autosize_columns(sheet)


def _build_summary_sheet(sheet, parsed_data: ParsedOFXData) -> None:
    title_fill = PatternFill(fill_type="solid", fgColor=NAVY)
    accent_fill = PatternFill(fill_type="solid", fgColor=SUMMARY_BG)
    white_font = Font(color=WHITE, bold=True, size=14)
    label_font = Font(color=NAVY_LIGHT, bold=True)
    value_font = Font(color=TEXT, bold=True)
    border = Border(
        left=Side(style="thin", color=BORDER),
        right=Side(style="thin", color=BORDER),
        top=Side(style="thin", color=BORDER),
        bottom=Side(style="thin", color=BORDER),
    )

    sheet.merge_cells("A1:B1")
    title_cell = sheet["A1"]
    title_cell.value = "Resumo da Conversão"
    title_cell.fill = title_fill
    title_cell.font = white_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    rows = [
        ("Arquivo original", parsed_data.original_filename),
        ("Banco identificado", parsed_data.banco),
        ("Data e hora da conversão", parsed_data.processed_at.strftime("%d/%m/%Y %H:%M:%S")),
        ("Total de entradas", float(parsed_data.summary.total_entradas)),
        ("Total de saídas", float(parsed_data.summary.total_saidas)),
        ("Saldo final", float(parsed_data.summary.saldo)),
        ("Quantidade de movimentações", parsed_data.summary.quantidade_movimentacoes),
    ]

    for index, (label, value) in enumerate(rows, start=3):
        label_cell = sheet.cell(row=index, column=1, value=label)
        value_cell = sheet.cell(row=index, column=2, value=value)
        label_cell.font = label_font
        value_cell.font = value_font
        label_cell.fill = accent_fill
        value_cell.fill = accent_fill
        label_cell.border = border
        value_cell.border = border
        label_cell.alignment = Alignment(horizontal="left")
        value_cell.alignment = Alignment(horizontal="left")

        if label in {"Total de entradas", "Total de saídas", "Saldo final"}:
            value_cell.number_format = 'R$ #,##0.00'

    _autosize_columns(sheet)


def _style_header_row(sheet, row_number: int) -> None:
    fill = PatternFill(fill_type="solid", fgColor=NAVY)
    font = Font(color=WHITE, bold=True)
    border = Border(bottom=Side(style="medium", color=CYAN))

    for cell in sheet[row_number]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def _autosize_columns(sheet) -> None:
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            length = len(str(value))
            if length > max_length:
                max_length = length
        sheet.column_dimensions[column_letter].width = min(max_length + 4, 40)


def _to_float(value: str) -> float | None:
    return float(value) if value else None


def _to_excel_date(value: str) -> datetime | None:
    if not value:
        return None
    return datetime.strptime(value, "%d/%m/%Y")
